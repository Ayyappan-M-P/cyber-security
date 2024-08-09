from flask import Flask, request
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'clicks.xlsx'

def save_to_excel(email):
    # Check if the Excel file already exists
    if not os.path.exists(EXCEL_FILE):
        # Create a new workbook and add a header
        wb = Workbook()
        ws = wb.active
        ws.append(['Email', 'Timestamp'])
        wb.save(EXCEL_FILE)

    try:
        # Load the existing workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Append the email and current timestamp
        ws.append([email, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error writing to Excel file: {e}")

@app.route('/track')
def track():
    email = request.args.get('email')
    if email:
        save_to_excel(email)
        return f"Email {email} has been tracked.", 200
    return "No email provided.", 400

if __name__ == '__main__':
    app.run(port=5000, debug=True)
